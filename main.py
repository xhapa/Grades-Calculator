from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.styles import colors

def close_xlsx(workbook):
    try:
        workbook.save('scores_table.xlsx')
        print('Is`t OK !')
    except Exception as e:
        print(e)

def fetch_data(sheet):      
    num_students = int(sheet['A1'].value)
    grades_percentage = [sheet['2'][i].value for i in range(len(sheet['2']))]
    grade_type = [sheet['1'][i].value for i in range(1, len(sheet['1']))]
    students_grades = {sheet[f'{row}'][0].value:{grade_type[col]: sheet[f'{row}'][col+1].value for col in range(len(grade_type))} for row in range(4, num_students+4)}

    return num_students, grades_percentage, students_grades, grade_type

def calculate_final_grade(students_grades, grade_percentage, grade_type):
    final_grades = {}
    for student in students_grades:
        final_score = 0
        for i in range(1, len(grade_percentage)):
            final_score += float(students_grades[student][grade_type[i-1]])*grade_percentage[i]
        final_grades[student] = round(final_score,3)

    return final_grades

def fill_sheet(o_sheet, final_grades):
    names = list(final_grades.keys())
    grade = list(final_grades.values())
    red_fill = PatternFill(start_color='FFDCD6',
                   end_color='FFDCD6',
                   fill_type='solid')
    green_fill = PatternFill(start_color='D9FFD6',
                   end_color='D9FFD6',
                   fill_type='solid')
    blue_fill = PatternFill(start_color='D6EDFF',
                   end_color='D6EDFF',
                   fill_type='solid')

    o_sheet['A1'] = 'Students'
    o_sheet['B1'] = 'Final Grades'
    o_sheet['A1'].fill = blue_fill
    o_sheet['B1'].fill = blue_fill

    for i in range(1,len(final_grades)+1):
        o_sheet['A'+f'{i+1}'].value = names[i-1]
        o_sheet['B'+f'{i+1}'].value = grade[i-1]
        if grade[i-1] < 3.0:
            o_sheet['B'+f'{i+1}'].fill = red_fill
        else:
            o_sheet['B'+f'{i+1}'].fill = green_fill

def main():
    workbook= load_workbook('scores_table.xlsx')

    if 'Final Grades' not in workbook.sheetnames:
        workbook.create_sheet('Final Grades')
    

    num_students, grades_percent, students_grades, grade_type= fetch_data(workbook['Grades'])
    final_grades = calculate_final_grade(students_grades, grades_percent, grade_type)
    fill_sheet(workbook['Final Grades'], final_grades)

    close_xlsx(workbook)

if __name__ == '__main__':
    main()